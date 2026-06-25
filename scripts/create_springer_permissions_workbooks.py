from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PACKAGE = ROOT / "output/springer_submission/Data_Engineering_for_Large_Foundation_Models_A_Handbook"
PERMISSIONS_DIR = PACKAGE / "03_Third_Party_Permissions"
PRINT_MANIFEST = PACKAGE / "01_Source_Files/Figures_Print_Formats/figures_print_format_manifest.csv"
ALT_TEXT_XLSX = PACKAGE / "01_Source_Files/Accessibility/springer_alt_text_inventory.xlsx"
CHAPTER_MANIFEST = PACKAGE / "01_Source_Files/LaTeX/chapter_tex_manifest.csv"
DOCS_EN = ROOT / "docs/en"

TODAY = date.today().isoformat()

IMAGE_RE = re.compile(r"!\[([^\]]*)\]\(([^)]+)\)")
LINK_RE = re.compile(r"(?<!!)\[([^\]]+)\]\(([^)]+)\)")
REF_HEADING_RE = re.compile(r"^#{2,6}\s+(References|Bibliography)\s*$", re.IGNORECASE)
FIG_LABEL_RE = re.compile(r"(?:Fig(?:ure)?|图)\s*([A-Z]?\d+(?:[-.]\d+)?|P\d+-\d+)", re.IGNORECASE)
DATASET_TERMS = re.compile(
    r"\b(?:dataset|benchmark|corpus|model card|data card|Common Crawl|LAION|FineWeb|Dolma|DataComp|Pexels|Hugging Face|GitHub|Ray|MLflow|DVC|Kubernetes|vLLM|OWASP|GDPR|NIST|ISO/IEC)\b",
    re.IGNORECASE,
)


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def sha256_file(path: Path) -> str:
    if not path.exists() or not path.is_file():
        return ""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def chapter_meta() -> dict[str, dict[str, str]]:
    rows = read_csv_rows(CHAPTER_MANIFEST)
    return {row.get("source_markdown", ""): row for row in rows}


def print_meta() -> dict[str, dict[str, str]]:
    rows = read_csv_rows(PRINT_MANIFEST)
    out = {}
    for row in rows:
        key = row.get("original_image_path", "")
        if key:
            out[key] = row
    return out


def alt_text_meta() -> dict[str, dict[str, str]]:
    if not ALT_TEXT_XLSX.exists():
        return {}
    wb = load_workbook(ALT_TEXT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        key = ""
        for candidate in ("image_path", "Image Path", "path", "figure_file", "file"):
            if candidate in item and item[candidate]:
                key = str(item[candidate])
                break
        if key:
            rows[key] = {k: "" if v is None else str(v) for k, v in item.items()}
    return rows


def iter_en_markdown() -> list[Path]:
    return sorted(p for p in DOCS_EN.rglob("*.md") if p.name not in {"translation-status.md", "translation-style-guide.md"})


def normalize_target(source: Path, target: str) -> tuple[str, Path]:
    clean = target.split("#", 1)[0].strip()
    if re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*:", clean):
        return clean, Path(clean)
    abs_path = (source.parent / clean).resolve()
    try:
        rel = abs_path.relative_to(ROOT)
        return rel.as_posix(), abs_path
    except ValueError:
        return str(abs_path), abs_path


def line_number(text: str, pos: int) -> int:
    return text[:pos].count("\n") + 1


def guess_item_id(path_or_text: str, source: str, seq: int, prefix: str) -> str:
    m = FIG_LABEL_RE.search(path_or_text)
    if m:
        label = m.group(1).replace(".", "-")
        return f"{prefix}-{label}"
    base = Path(source).stem.replace("_", "-")
    return f"{prefix}-{base}-{seq:03d}"


def extract_tables(text: str) -> list[tuple[int, int, str]]:
    lines = text.splitlines()
    tables = []
    i = 0
    while i < len(lines) - 1:
        if "|" in lines[i] and "|" in lines[i + 1] and re.search(r"\|\s*:?-{3,}:?\s*\|", lines[i + 1]):
            start = i
            j = i + 2
            while j < len(lines) and "|" in lines[j] and lines[j].strip():
                j += 1
            snippet = "\n".join(lines[start : min(j, start + 4)])
            tables.append((start + 1, j, snippet[:500]))
            i = j
        else:
            i += 1
    return tables


def extract_code_blocks(text: str) -> list[tuple[int, int, str]]:
    lines = text.splitlines()
    out = []
    i = 0
    while i < len(lines):
        if lines[i].startswith("```") or lines[i].startswith("~~~"):
            fence = lines[i][:3]
            start = i
            lang = lines[i][3:].strip()
            i += 1
            while i < len(lines) and not lines[i].startswith(fence):
                i += 1
            end = i if i < len(lines) else len(lines) - 1
            if end - start >= 8:
                out.append((start + 1, end + 1, lang))
        i += 1
    return out


def figure_rows(meta_by_chapter: dict[str, dict[str, str]], pm: dict[str, dict[str, str]], am: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    rows = []
    seq = 1
    for md in iter_en_markdown():
        rel_md = md.relative_to(DOCS_EN).as_posix()
        text = md.read_text(encoding="utf-8")
        for match in IMAGE_RE.finditer(text):
            alt, target = match.group(1).strip(), match.group(2).strip()
            norm, abs_path = normalize_target(md, target)
            pmeta = pm.get(norm, {})
            altmeta = am.get(norm, {})
            item_id = guess_item_id(norm + " " + alt, rel_md, seq, "FIG")
            chapter = meta_by_chapter.get(rel_md, {})
            rows.append(
                {
                    "Item ID": item_id,
                    "Item Type": "Figure/Image",
                    "Book Location": chapter.get("title", rel_md),
                    "Source Markdown": rel_md,
                    "Line": str(line_number(text, match.start())),
                    "Caption/Alt Text": alt,
                    "Original File/Path": norm,
                    "Print File/Path": pmeta.get("print_format_path", ""),
                    "Original Format": pmeta.get("original_format", abs_path.suffix.lstrip(".").lower()),
                    "Print Format": pmeta.get("print_format", ""),
                    "SHA256": pmeta.get("original_sha256", sha256_file(abs_path)),
                    "Origin/Provenance": "Author-created or author-redrawn explanatory figure; final editor signoff states manuscript figures were manually redrawn and reviewed.",
                    "Third-Party Source Basis": "None identified in submitted figure file; surrounding text/caption may cite external methods, datasets, tools, or papers summarized by this original diagram.",
                    "Permission Required?": "No, unless later review identifies reproduced/adapted third-party visual content.",
                    "License/Permission Evidence": "Originality and permissions declaration; figure rights signoff; print-format manifest.",
                    "Credit Line / Caption Note": "Use manuscript caption; cite external concepts in surrounding text or references where applicable.",
                    "AI-Generated Image?": "No per declaration.",
                    "Accessibility/Alt Text Status": "Covered in alt-text inventory" if altmeta else "Check alt-text inventory",
                    "Print-Format Status": pmeta.get("status", "not found in print manifest"),
                    "Review Status": "Cleared by declaration; keep evidence with submission package.",
                    "Reviewer/Owner": chapter.get("authors", "Jun Yu"),
                    "Notes": pmeta.get("message", ""),
                }
            )
            seq += 1
    return rows


def table_rows(meta_by_chapter: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    rows = []
    seq = 1
    for md in iter_en_markdown():
        rel_md = md.relative_to(DOCS_EN).as_posix()
        text = md.read_text(encoding="utf-8")
        chapter = meta_by_chapter.get(rel_md, {})
        for start, end, snippet in extract_tables(text):
            rows.append(
                {
                    "Item ID": f"TAB-{Path(rel_md).stem.replace('_', '-')}-{seq:03d}",
                    "Item Type": "Table",
                    "Book Location": chapter.get("title", rel_md),
                    "Source Markdown": rel_md,
                    "Line": str(start),
                    "Caption/Alt Text": snippet.replace("\n", " ")[:320],
                    "Original File/Path": rel_md,
                    "Print File/Path": "",
                    "Original Format": "Markdown table",
                    "Print Format": "LaTeX/PDF table",
                    "SHA256": hashlib.sha256(snippet.encode("utf-8")).hexdigest(),
                    "Origin/Provenance": "Author-created technical summary, comparison table, checklist, or template unless a later review identifies reproduced/adapted third-party tabular content.",
                    "Third-Party Source Basis": "If summarizing external tools, datasets, papers, standards, or policies, verify citation in local text/references.",
                    "Permission Required?": "No for author-created tables; Yes if copied or closely adapted from third-party table.",
                    "License/Permission Evidence": "Originality and permissions declaration; chapter references; add source snapshot if a table is adapted from a third-party source.",
                    "Credit Line / Caption Note": "Use manuscript table title/note; add 'Adapted from...' only when applicable and permission/license supports it.",
                    "AI-Generated Image?": "N/A",
                    "Accessibility/Alt Text Status": "Text table; ensure table header structure remains editable.",
                    "Print-Format Status": "Generated from manuscript source",
                    "Review Status": "Requires final editorial spot-check for copied/adapted tables.",
                    "Reviewer/Owner": chapter.get("authors", "Jun Yu"),
                    "Notes": f"Rows {start}-{end}",
                }
            )
            seq += 1
    return rows


def permissions_rows(figs: list[dict[str, str]], tabs: list[dict[str, str]], meta_by_chapter: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    rows = []
    for item in figs + tabs:
        rows.append(
            {
                "Permission Item ID": item["Item ID"],
                "Content Type": item["Item Type"],
                "Chapter/Location": item["Book Location"],
                "Source File": item["Source Markdown"],
                "Line/Page": item["Line"],
                "Description": item["Caption/Alt Text"],
                "Third-Party Material?": "No known third-party reproduced content",
                "Original / Adapted / Reproduced": "Original/author-redrawn" if item["Item Type"].startswith("Figure") else "Author-created",
                "Source / Rights Holder": "Author team",
                "License / Permission Basis": item["License/Permission Evidence"],
                "Permission Required?": item["Permission Required?"],
                "Permission Status": "Cleared by declaration / no separate permission currently known",
                "Evidence File or URL": "originality_and_permissions_declaration.md; _Internal_Not_For_Submission/Audit_Reports/figure_rights_signoff.md",
                "Required Credit Line": item["Credit Line / Caption Note"],
                "Springer Risk Level": "Low if provenance statement remains true; Medium if adapted source later identified",
                "Action Needed": "Keep in checklist; add evidence if Springer requests item-level proof.",
                "Owner": item["Reviewer/Owner"],
                "Review Date": TODAY,
                "Notes": item["Notes"],
            }
        )

    code_seq = 1
    data_seq = 1
    quote_seq = 1
    for md in iter_en_markdown():
        rel_md = md.relative_to(DOCS_EN).as_posix()
        text = md.read_text(encoding="utf-8")
        chapter = meta_by_chapter.get(rel_md, {})
        location = chapter.get("title", rel_md)
        owner = chapter.get("authors", "Jun Yu")

        for start, end, lang in extract_code_blocks(text):
            rows.append(
                {
                    "Permission Item ID": f"CODE-{Path(rel_md).stem.replace('_', '-')}-{code_seq:03d}",
                    "Content Type": "Code block",
                    "Chapter/Location": location,
                    "Source File": rel_md,
                    "Line/Page": str(start),
                    "Description": f"Code block ({lang or 'plain'}) lines {start}-{end}",
                    "Third-Party Material?": "Likely author-created example; verify if copied from external project.",
                    "Original / Adapted / Reproduced": "Author-created example unless otherwise noted",
                    "Source / Rights Holder": "Author team / code repository contributors if copied",
                    "License / Permission Basis": "Manuscript declaration; repository license for book examples. Add upstream license evidence if code is copied/adapted.",
                    "Permission Required?": "No for original examples; verify for copied external code.",
                    "Permission Status": "Needs spot-check if block is copied from external repository",
                    "Evidence File or URL": "",
                    "Required Credit Line": "Credit external source in text/comment if copied/adapted.",
                    "Springer Risk Level": "Medium",
                    "Action Needed": "Spot-check long code block provenance.",
                    "Owner": owner,
                    "Review Date": TODAY,
                    "Notes": "",
                }
            )
            code_seq += 1

        in_refs = False
        for idx, line in enumerate(text.splitlines(), start=1):
            if REF_HEADING_RE.match(line.strip()):
                in_refs = True
                continue
            if not in_refs and DATASET_TERMS.search(line):
                rows.append(
                    {
                        "Permission Item ID": f"SRC-{Path(rel_md).stem.replace('_', '-')}-{data_seq:03d}",
                        "Content Type": "Dataset/tool/standard/source mention",
                        "Chapter/Location": location,
                        "Source File": rel_md,
                        "Line/Page": str(idx),
                        "Description": line.strip()[:500],
                        "Third-Party Material?": "Third-party name/concept cited; no reproduced content identified by this automated pass.",
                        "Original / Adapted / Reproduced": "Mention/citation/summary",
                        "Source / Rights Holder": "Named external project, publisher, standards body, dataset owner, or institution",
                        "License / Permission Basis": "Citation/reference; add license or source snapshot if content/data/figure is reproduced.",
                        "Permission Required?": "Usually no for nominative mention or brief technical summary; verify if data, images, tables, or substantial text are reproduced.",
                        "Permission Status": "Citation/source evidence to verify",
                        "Evidence File or URL": "See chapter references and reference confirmation workbook.",
                        "Required Credit Line": "Cite source in text/references; do not imply endorsement.",
                        "Springer Risk Level": "Medium",
                        "Action Needed": "Check whether mention is only descriptive or includes reproduced/adapted content.",
                        "Owner": owner,
                        "Review Date": TODAY,
                        "Notes": "",
                    }
                )
                data_seq += 1

            if line.count('"') >= 2 and len(line.strip()) > 180:
                rows.append(
                    {
                        "Permission Item ID": f"QUOTE-{Path(rel_md).stem.replace('_', '-')}-{quote_seq:03d}",
                        "Content Type": "Potential quotation",
                        "Chapter/Location": location,
                        "Source File": rel_md,
                        "Line/Page": str(idx),
                        "Description": line.strip()[:500],
                        "Third-Party Material?": "Potential quoted text; requires editorial review.",
                        "Original / Adapted / Reproduced": "Unknown",
                        "Source / Rights Holder": "To be confirmed",
                        "License / Permission Basis": "Citation or permission may be needed depending on length/substantiality.",
                        "Permission Required?": "Review required",
                        "Permission Status": "Open for spot-check",
                        "Evidence File or URL": "",
                        "Required Credit Line": "Add citation/quotation attribution if retained.",
                        "Springer Risk Level": "Medium",
                        "Action Needed": "Verify quotation length, attribution, and whether paraphrase is preferable.",
                        "Owner": owner,
                        "Review Date": TODAY,
                        "Notes": "",
                    }
                )
                quote_seq += 1

    return rows


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        values = [str(header)]
        for row_idx in range(2, min(ws.max_row, 80) + 1):
            values.append(str(ws.cell(row_idx, col_idx).value or ""))
        width = min(max(max(len(v) for v in values) + 2, 12), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def summary_rows(figs: list[dict[str, str]], tabs: list[dict[str, str]], perms: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        {"Metric": "Generated on", "Value": TODAY, "Notes": ""},
        {"Metric": "Figure/image rows", "Value": str(len(figs)), "Notes": "Extracted from docs/en Markdown image references and enriched with print-format manifest."},
        {"Metric": "Table rows", "Value": str(len(tabs)), "Notes": "Extracted from Markdown pipe tables."},
        {"Metric": "Permission checklist rows", "Value": str(len(perms)), "Notes": "Includes figures, tables, code blocks, source mentions, and potential quotations."},
        {"Metric": "Declaration basis", "Value": "originality_and_permissions_declaration.md", "Notes": "Does not replace Springer forms or item-specific permission letters if required."},
        {"Metric": "Residual action", "Value": "Spot-check medium-risk rows", "Notes": "Especially copied/adapted code, potential quotation lines, stock/video images, logos/screenshots, and external dataset visuals."},
    ]


def write_workbooks() -> tuple[Path, Path, int, int, int]:
    PERMISSIONS_DIR.mkdir(parents=True, exist_ok=True)
    meta = chapter_meta()
    pm = print_meta()
    am = alt_text_meta()
    figs = figure_rows(meta, pm, am)
    tabs = table_rows(meta)
    perms = permissions_rows(figs, tabs, meta)

    fig_headers = [
        "Item ID",
        "Item Type",
        "Book Location",
        "Source Markdown",
        "Line",
        "Caption/Alt Text",
        "Original File/Path",
        "Print File/Path",
        "Original Format",
        "Print Format",
        "SHA256",
        "Origin/Provenance",
        "Third-Party Source Basis",
        "Permission Required?",
        "License/Permission Evidence",
        "Credit Line / Caption Note",
        "AI-Generated Image?",
        "Accessibility/Alt Text Status",
        "Print-Format Status",
        "Review Status",
        "Reviewer/Owner",
        "Notes",
    ]
    perm_headers = [
        "Permission Item ID",
        "Content Type",
        "Chapter/Location",
        "Source File",
        "Line/Page",
        "Description",
        "Third-Party Material?",
        "Original / Adapted / Reproduced",
        "Source / Rights Holder",
        "License / Permission Basis",
        "Permission Required?",
        "Permission Status",
        "Evidence File or URL",
        "Required Credit Line",
        "Springer Risk Level",
        "Action Needed",
        "Owner",
        "Review Date",
        "Notes",
    ]
    summary_headers = ["Metric", "Value", "Notes"]

    fig_wb = Workbook()
    del fig_wb[fig_wb.sheetnames[0]]
    add_sheet(fig_wb, "Summary", summary_headers, summary_rows(figs, tabs, perms))
    add_sheet(fig_wb, "Figures", fig_headers, figs)
    add_sheet(fig_wb, "Tables", fig_headers, tabs)
    fig_path = PERMISSIONS_DIR / "figure_table_provenance_manifest.xlsx"
    fig_wb.save(fig_path)

    perm_wb = Workbook()
    del perm_wb[perm_wb.sheetnames[0]]
    add_sheet(perm_wb, "Summary", summary_headers, summary_rows(figs, tabs, perms))
    add_sheet(perm_wb, "Permissions Checklist", perm_headers, perms)
    legend_rows = [
        {"Field": "Permission Required?", "Meaning": "No = no separate permission known; Review required = editorial/legal check needed; Yes = permission/license evidence must be attached.", "Example": ""},
        {"Field": "Springer Risk Level", "Meaning": "Low/Medium/High editorial triage only, not legal advice.", "Example": ""},
        {"Field": "Evidence File or URL", "Meaning": "Point to declaration, signoff, permission letter, license snapshot, source page, or publisher correspondence.", "Example": "license_and_source_snapshots/example.pdf"},
        {"Field": "Credit Line", "Meaning": "Use exact credit line required by license/permission when material is adapted or reproduced.", "Example": "Adapted from X, licensed under CC BY 4.0."},
    ]
    add_sheet(perm_wb, "Field Guide", ["Field", "Meaning", "Example"], legend_rows)
    perm_path = PERMISSIONS_DIR / "third_party_permissions_checklist.xlsx"
    perm_wb.save(perm_path)
    return perm_path, fig_path, len(perms), len(figs), len(tabs)


def main() -> None:
    perm_path, fig_path, perm_count, fig_count, tab_count = write_workbooks()
    print(f"permissions={perm_path}")
    print(f"provenance={fig_path}")
    print(json.dumps({"permissions_rows": perm_count, "figure_rows": fig_count, "table_rows": tab_count}, ensure_ascii=False))


if __name__ == "__main__":
    main()

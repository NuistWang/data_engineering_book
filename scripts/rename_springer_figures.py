#!/usr/bin/env python3
"""Rename repository figure files to Springer-friendly names and update references."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
IMAGE_ROOTS = [ROOT / "docs" / "images", ROOT / "images"]
OUTPUT_PLAN = ROOT / "output" / "springer_submission" / "figure_rename_plan.csv"
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".svg", ".webp", ".gif"}
TEXT_EXTS = {".md", ".yml", ".yaml", ".py", ".csv", ".json", ".tex", ".html"}
SKIP_DIRS = {".git", "node_modules", "output", ".venv", "__pycache__", ".mypy_cache", ".pytest_cache"}
ACCESSIBILITY_DIR = ROOT / "publishing" / "accessibility"

AUTHOR_PREFIXES = {
    "part1/ch01_": "Jun Yu; Changwen Chen; Ke Wang",
    "part1/ch02_": "Jun Yu; Changwen Chen; Ke Wang",
    "part1/ch03_": "Jun Yu; Ke Wang; Changwen Chen",
    "part2/ch04_": "Jun Yu; Ke Wang; Changwen Chen",
    "part2/ch05_": "Jun Yu; Ke Wang; Changwen Chen",
    "part2/ch06_": "Ke Wang; Fan Yu; Jun Yu",
    "part2/ch07_": "Ke Wang; Fan Yu; Jun Yu",
    "part3/ch08_": "Jun Yu; Ke Wang; Cong Wang",
    "part3/ch09_": "Jun Yu; Ke Wang; Cong Wang",
    "part3/ch10_": "Ke Wang; Cong Wang; Jun Yu",
    "part3/ch11_": "Ke Wang; Cong Wang; Jun Yu",
    "part4/ch12_": "Jun Yu; Ran Zhang; Yang Luo",
    "part4/ch13_": "Jun Yu; Ran Zhang; Yang Luo",
    "part4/ch14_": "Ran Zhang; Yang Luo; Jun Yu",
    "part5/ch15_": "Cong Wang; Ran Zhang; Jun Yu",
    "part5/ch16_": "Cong Wang; Ran Zhang; Jun Yu",
    "part5/ch17_": "Ran Zhang; Yang Luo; Jun Yu",
    "part6/ch18_": "Jun Yu; Ran Zhang; Zhongyi Liu",
    "part6/ch19_": "Jun Yu; Ran Zhang; Zhongyi Liu",
    "part6/ch20_": "Ran Zhang; Zhongyi Liu; Jun Yu",
    "part7/ch21_": "Wenzhuo Du; Gongpeng Zhao; Jun Yu",
    "part7/ch22_": "Wenzhuo Du; Gongpeng Zhao; Jun Yu",
    "part7/ch23_": "Jun Yu; Wenzhuo Du; Gongpeng Zhao",
    "part8/ch24_": "Jun Yu; Wenzhuo Du; Can Wang",
    "part8/ch25_": "Wenzhuo Du; Can Wang; Jun Yu",
    "part8/ch26_": "Wenzhuo Du; Can Wang; Jun Yu",
    "part9/ch27_": "Ran Zhang; Feng Zhao; Wenzhuo Du",
    "part9/ch28_": "Zhongyi Liu; Ye Yu; Wenzhuo Du",
    "part9/ch29_": "Zhongyi Liu; Wenzhuo Du; Jun Yu",
    "part9/ch30_": "Yang Luo; Fang Gao; Wenzhuo Du",
    "part10/ch31_": "Jun Yu; Zhili Wang; Zhongyi Liu",
    "part10/ch32_": "Jun Yu; Zhili Wang; Zhongyi Liu",
    "part10/ch33_": "Zhili Wang; Zhongyi Liu; Jun Yu",
    "part10/ch34_": "Yang Luo; Zhili Wang; Jun Yu",
    "part10/ch35_": "Yang Luo; Zhili Wang; Jun Yu",
    "part11/ch36_": "Zhili Wang; Xin Xu; Jun Yu",
    "part11/ch37_": "Zhili Wang; Xin Xu; Jun Yu",
    "part12/ch38_": "Guanlin Mu; Xuhong Cao",
    "part12/ch39_": "Guanlin Mu; Xuhong Cao",
    "part12/ch40_": "Guanjun Liu; Yuefeng Zou",
    "part12/ch41_": "Lin Xu; Xinyu Chen",
    "part12/ch42_": "Fengxin Chen; Xuan Li",
    "part12/ch43_": "Xuan Li; Fengxin Chen",
    "part13/ch44_": "Ke Wang; Jiaen Liang; Jun Yu",
    "part13/ch45_": "Cong Wang; Xin Xu; Wei Huang",
    "part13/ch46_": "Xin Xu; Shengping Liu; Fan Yu",
    "part13/ch47_": "Xuhong Cao; Ke Wang; Qingsong Liu",
    "part13/ch48_": "Ran Zhang; Jianqing Sun; Fan Yu",
    "part14/p01_": "Xin Xu; Ran Zhang; Jun Yu",
    "part14/p02_": "Xin Xu; Ran Zhang; Jun Yu",
    "part14/p03_": "Jun Yu; Xin Xu; Wenzhuo Du",
    "part14/p04_": "Xin Xu; Wenzhuo Du; Jun Yu",
    "part14/p05_": "Xuhong Cao; Ke Wang; Jun Yu",
    "part14/p06_": "Cong Wang; Xin Xu; Ke Wang",
    "part14/p07_": "Jun Yu; Xin Xu; Zhili Wang",
    "part14/p08_": "Jun Yu; Xin Xu; Zhili Wang",
    "part14/p09_": "Zhongyi Liu; Xin Xu; Guanlin Mu",
    "part14/p10_": "Ke Wang; Xin Xu; Guanlin Mu",
    "part14/p11_": "Jun Yu; Ke Wang; Yang Luo",
    "part14/p12_": "Cong Wang; Xin Xu; Yang Luo",
    "part14/p13_": "Jun Yu; Ke Wang; Wenzhuo Du",
    "part14/p14_": "Yang Luo; Ran Zhang; Wenzhuo Du",
    "part14/p15_": "Xuhong Cao; Zhongyi Liu; Jun Yu",
}


@dataclass
class ImageRef:
    source: Path
    alt: str
    line: int


@dataclass
class RenameRow:
    old_path: Path
    new_path: Path
    unit_label: str
    figure_label: str
    authors: str
    referenced: bool


def safe_slug(value: str, *, max_len: int = 80) -> str:
    value = value.replace("&", " and ")
    value = re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-")
    value = re.sub(r"-{2,}", "-", value)
    return value[:max_len].strip("-") or "untitled"


def first_author_surname(authors: str) -> str:
    first = re.split(r";|,", authors, maxsplit=1)[0].strip()
    return safe_slug(first.split()[-1], max_len=24) if first else "Yu"


def authors_for_unit(unit: str) -> str:
    for prefix, authors in AUTHOR_PREFIXES.items():
        if unit.startswith(prefix):
            return authors
    return "Jun Yu"


def iter_image_files() -> list[Path]:
    return sorted(
        path
        for root in IMAGE_ROOTS
        if root.exists()
        for path in root.rglob("*")
        if path.is_file() and path.suffix.lower() in IMAGE_EXTS
    )


def iter_text_files() -> list[Path]:
    files: list[Path] = []
    for path in ROOT.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in TEXT_EXTS:
            continue
        rel_parts = path.relative_to(ROOT).parts
        if any(part in SKIP_DIRS for part in rel_parts):
            continue
        if rel_parts[:2] == ("publishing", "pdf_preview"):
            continue
        if rel_parts[:2] == ("publishing", "final_review"):
            continue
        files.append(path)
    return sorted(files)


def markdown_refs(path: Path) -> list[tuple[str, str, int]]:
    refs: list[tuple[str, str, int]] = []
    text = path.read_text(encoding="utf-8", errors="replace")
    for line_no, line in enumerate(text.splitlines(), start=1):
        for match in re.finditer(r"!\[([^\]]*)]\(([^)]+)\)", line):
            refs.append((match.group(2), match.group(1), line_no))
        for match in re.finditer(r"<img[^>]+src=[\"']([^\"']+)[\"'][^>]*>", line, flags=re.I):
            refs.append((match.group(1), "", line_no))
    return refs


def resolve_ref(source: Path, raw: str) -> Path | None:
    url = raw.strip().split("#", 1)[0].split("?", 1)[0]
    if not url or re.match(r"^(?:https?:|data:|file:|#)", url):
        return None
    path = (source.parent / url).resolve()
    if path.exists() and path.is_file() and path.is_relative_to(ROOT):
        return path
    return None


def reference_map() -> dict[Path, list[ImageRef]]:
    refs: dict[Path, list[ImageRef]] = {}
    candidates = [ROOT / "README.md", *sorted((ROOT / "docs").rglob("*.md"))]
    for source in candidates:
        if not source.exists():
            continue
        for raw, alt, line in markdown_refs(source):
            resolved = resolve_ref(source, raw)
            if resolved is None or resolved.suffix.lower() not in IMAGE_EXTS:
                continue
            refs.setdefault(resolved, []).append(ImageRef(source=source, alt=alt, line=line))
    return refs


def source_unit(source: Path) -> str:
    rel = source.relative_to(ROOT).as_posix()
    if rel == "README.md":
        return "book"
    if rel.startswith("docs/en/"):
        return rel.removeprefix("docs/en/")
    if rel.startswith("docs/zh/"):
        return rel.removeprefix("docs/zh/")
    return rel


def unit_from_path(path: Path, refs: list[ImageRef]) -> str:
    explicit = explicit_unit_from_image_name(path)
    if explicit:
        return explicit
    if refs:
        en_refs = [ref for ref in refs if "/docs/en/" in ref.source.as_posix()]
        zh_refs = [ref for ref in refs if "/docs/zh/" in ref.source.as_posix()]
        chosen = en_refs[0] if en_refs else (zh_refs[0] if zh_refs else refs[0])
        return source_unit(chosen.source)
    rel = path.relative_to(ROOT).as_posix()
    stem = path.stem
    match = re.search(r"ch(\d{1,2})_", stem, flags=re.I)
    if match:
        chapter = int(match.group(1))
        part = part_for_chapter(chapter)
        return f"part{part}/ch{chapter:02d}_"
    match = re.search(r"图(\d{1,2})[_-]", stem)
    if match:
        chapter = int(match.group(1))
        part = part_for_chapter(chapter)
        return f"part{part}/ch{chapter:02d}_"
    match = re.search(r"p(\d{1,2})[_/-]", rel, flags=re.I)
    if match:
        project = int(match.group(1))
        return f"part14/p{project:02d}_"
    if rel.startswith("docs/images/book_structure"):
        return "front_matter_guide.md"
    if rel.startswith("images/structure"):
        return "book"
    return rel


def explicit_unit_from_image_name(path: Path) -> str | None:
    rel = path.relative_to(ROOT).as_posix()
    stem = path.stem
    match = re.search(r"ch(\d{1,2})_", stem, flags=re.I)
    if match:
        chapter = int(match.group(1))
        return f"part{part_for_chapter(chapter)}/ch{chapter:02d}_"
    match = re.search(r"[图圖図](\d{1,2})[_-]", stem)
    if match:
        chapter = int(match.group(1))
        return f"part{part_for_chapter(chapter)}/ch{chapter:02d}_"
    match = re.search(r"(?:^|/)p(\d{1,2})(?:[_/-]|$)", rel, flags=re.I)
    if match:
        return f"part14/p{int(match.group(1)):02d}_"
    match = re.search(r"p(\d{1,2})_", stem, flags=re.I)
    if match:
        return f"part14/p{int(match.group(1)):02d}_"
    return None


def part_for_chapter(chapter: int) -> int:
    ranges = [
        (1, 3, 1),
        (4, 7, 2),
        (8, 11, 3),
        (12, 14, 4),
        (15, 17, 5),
        (18, 20, 6),
        (21, 23, 7),
        (24, 26, 8),
        (27, 30, 9),
        (31, 35, 10),
        (36, 37, 11),
        (38, 43, 12),
        (44, 48, 13),
    ]
    for start, end, part in ranges:
        if start <= chapter <= end:
            return part
    return 1


def unit_label(unit: str) -> str:
    if unit == "book":
        return "Book"
    if "front_matter" in unit:
        return "FrontMatter"
    match = re.search(r"ch(\d{1,2})", unit, flags=re.I)
    if match:
        return f"Chap{int(match.group(1)):02d}"
    match = re.search(r"p(\d{1,2})", unit, flags=re.I)
    if match:
        return f"Project{int(match.group(1)):02d}"
    match = re.search(r"appendix_([a-h])", unit, flags=re.I)
    if match:
        return f"Appendix{match.group(1).upper()}"
    return "Book"


def figure_number(path: Path, refs: list[ImageRef], unit: str, fallback: int) -> tuple[str, str]:
    texts = [ref.alt for ref in refs if ref.alt]
    stem = path.stem
    label = unit_label(unit)
    for text in texts:
        match = re.search(r"(?:Figure|Fig\.?|图|圖|図)\s*(?:P)?\d{1,2}[-._](\d{1,2})(?:[-._]([A-Za-z0-9]+))?", text, flags=re.I)
        if match:
            suffix = panel_suffix(match.group(2) or "")
            return f"Fig{int(match.group(1)):02d}{suffix}", label
        match = re.search(r"(?:Figure|Fig\.?)\s*P(\d{1,2})[-._](\d{1,2})", text, flags=re.I)
        if match:
            return f"Fig{int(match.group(2)):02d}", label
    patterns = [
        r"ch\d{1,2}[_-](\d{1,2})(?:[^0-9]+(\d{1,2}))?$",
        r"ch\d{1,2}[_-](\d{1,2})",
        r"[图圖図]\d{1,2}[_-](\d{1,2})",
        r"p\d{1,2}[_-](\d{1,2})(?:[^0-9]+(\d{1,2}))?$",
        r"p\d{1,2}[_-](\d{1,2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, stem, flags=re.I)
        if match:
            suffix = panel_suffix(match.group(2) if len(match.groups()) >= 2 and match.group(2) else "")
            return f"Fig{int(match.group(1)):02d}{suffix}", label
    return f"Fig{fallback:02d}", label


def panel_suffix(value: str) -> str:
    if not value:
        return ""
    if value.isdigit():
        return f"-Panel{int(value):02d}"
    return f"-{safe_slug(value, max_len=12)}"


def language_suffix(path: Path, refs: list[ImageRef]) -> str:
    name = path.name
    if re.search(r"_en(?:\.|_)", name, flags=re.I) or name.endswith("_en.svg") or name.endswith("_en.png"):
        return "-EN"
    if re.search(r"[\u4e00-\u9fff]", name):
        return "-ZH"
    if refs:
        has_en = any("/docs/en/" in ref.source.as_posix() for ref in refs)
        has_zh = any("/docs/zh/" in ref.source.as_posix() for ref in refs)
        if has_zh and not has_en:
            return "-ZH"
    return ""


def build_plan() -> list[RenameRow]:
    refs = reference_map()
    seq_by_unit: dict[str, int] = {}
    raw_rows: list[tuple[Path, str, str, str, str, bool]] = []
    for path in iter_image_files():
        image_refs = refs.get(path, [])
        unit = unit_from_path(path, image_refs)
        seq_by_unit[unit] = seq_by_unit.get(unit, 0) + 1
        fig, label = figure_number(path, image_refs, unit, seq_by_unit[unit])
        authors = authors_for_unit(unit)
        suffix = language_suffix(path, image_refs)
        raw_rows.append((path, label, fig, authors, suffix, bool(image_refs)))

    target_counts: dict[Path, int] = {}
    proposed: list[RenameRow] = []
    for path, label, fig, authors, suffix, referenced in raw_rows:
        stem = f"{first_author_surname(authors)}-{label}-{fig}{suffix}"
        target = path.with_name(f"{stem}{path.suffix.lower()}")
        target_counts[target] = target_counts.get(target, 0) + 1
        proposed.append(RenameRow(path, target, label, fig, authors, referenced))

    target_seen: dict[Path, int] = {}
    final_rows: list[RenameRow] = []
    for row in proposed:
        target = row.new_path
        if target_counts[target] > 1:
            target_seen[target] = target_seen.get(target, 0) + 1
            target = target.with_name(f"{target.stem}-Alt{target_seen[target]:02d}{target.suffix}")
        final_rows.append(RenameRow(row.old_path, target, row.unit_label, row.figure_label, row.authors, row.referenced))
    return final_rows


def write_plan(rows: list[RenameRow]) -> None:
    OUTPUT_PLAN.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PLAN.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["old_path", "new_path", "unit_label", "figure_label", "authors", "referenced"],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "old_path": row.old_path.relative_to(ROOT).as_posix(),
                    "new_path": row.new_path.relative_to(ROOT).as_posix(),
                    "unit_label": row.unit_label,
                    "figure_label": row.figure_label,
                    "authors": row.authors,
                    "referenced": "yes" if row.referenced else "no",
                }
            )


def check_plan(rows: list[RenameRow]) -> None:
    targets = [row.new_path for row in rows]
    duplicate_targets = sorted({path for path in targets if targets.count(path) > 1})
    if duplicate_targets:
        raise RuntimeError("duplicate rename targets: " + ", ".join(path.relative_to(ROOT).as_posix() for path in duplicate_targets[:20]))
    sources = {row.old_path for row in rows}
    external_existing = [row.new_path for row in rows if row.new_path.exists() and row.new_path not in sources and row.new_path != row.old_path]
    if external_existing:
        raise RuntimeError("target already exists outside rename set: " + ", ".join(path.relative_to(ROOT).as_posix() for path in external_existing[:20]))


def replace_text_references(rows: list[RenameRow]) -> None:
    replacements: dict[str, str] = {}
    for row in rows:
        old_rel = row.old_path.relative_to(ROOT).as_posix()
        new_rel = row.new_path.relative_to(ROOT).as_posix()
        path_pairs = {
            old_rel: new_rel,
            old_rel.removeprefix("docs/"): new_rel.removeprefix("docs/"),
            "../" + old_rel.removeprefix("docs/"): "../" + new_rel.removeprefix("docs/"),
            "../../" + old_rel.removeprefix("docs/"): "../../" + new_rel.removeprefix("docs/"),
            "../../../" + old_rel.removeprefix("docs/"): "../../../" + new_rel.removeprefix("docs/"),
            old_rel.removeprefix("docs/images/"): new_rel.removeprefix("docs/images/"),
        }
        for old, new in path_pairs.items():
            if old and old != new:
                replacements[old] = new
    for path in iter_text_files():
        text = path.read_text(encoding="utf-8", errors="replace")
        new_text = text
        for old, new in sorted(replacements.items(), key=lambda item: len(item[0]), reverse=True):
            new_text = new_text.replace(old, new)
        if new_text != text:
            path.write_text(new_text, encoding="utf-8")


def replace_accessibility_workbook(rows: list[RenameRow]) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        print(f"[warn] openpyxl unavailable; skipped xlsx update: {exc}", file=sys.stderr)
        return
    xlsx = ACCESSIBILITY_DIR / "springer_alt_text_inventory.xlsx"
    if not xlsx.exists():
        return
    replacements = {row.old_path.relative_to(ROOT).as_posix(): row.new_path.relative_to(ROOT).as_posix() for row in rows}
    wb = load_workbook(xlsx)
    changed = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = cell.value
                new_value = value
                for old, new in sorted(replacements.items(), key=lambda item: len(item[0]), reverse=True):
                    new_value = new_value.replace(old, new)
                if new_value != value:
                    cell.value = new_value
                    changed = True
    if changed:
        wb.save(xlsx)


def rename_files(rows: list[RenameRow]) -> None:
    tmp_suffix = ".springer-rename-tmp"
    temp_pairs: list[tuple[Path, Path]] = []
    for row in rows:
        if row.old_path == row.new_path:
            continue
        tmp = row.old_path.with_name(row.old_path.name + tmp_suffix)
        if tmp.exists():
            tmp.unlink()
        row.old_path.rename(tmp)
        temp_pairs.append((tmp, row.new_path))
    for tmp, target in temp_pairs:
        target.parent.mkdir(parents=True, exist_ok=True)
        tmp.rename(target)


def refresh_accessibility_sidecars() -> None:
    csv_path = ACCESSIBILITY_DIR / "springer_alt_text_inventory.csv"
    json_path = ACCESSIBILITY_DIR / "springer_alt_text_inventory.json"
    xlsx_path = ACCESSIBILITY_DIR / "springer_alt_text_inventory.xlsx"
    if not xlsx_path.exists():
        return
    try:
        from openpyxl import load_workbook
    except Exception:
        return
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows([[cell_value(value) for value in row] for row in rows])
    headers = [str(value) if value is not None else "" for value in rows[1 if len(rows) > 1 else 0]]
    start = 2 if len(rows) > 1 else 1
    payload = []
    for row in rows[start:]:
        record = {}
        for key, value in zip(headers, row):
            if key:
                record[key] = cell_value(value)
        if record:
            payload.append(record)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def cell_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="Apply the rename plan and update references.")
    args = parser.parse_args()
    rows = build_plan()
    check_plan(rows)
    write_plan(rows)
    print(f"[ok] rename plan: {OUTPUT_PLAN}")
    print(f"[ok] images planned: {len(rows)}")
    print(f"[ok] referenced images: {sum(1 for row in rows if row.referenced)}")
    if not args.apply:
        print("[dry-run] pass --apply to rename files and update references")
        return 0
    backup = OUTPUT_PLAN.with_suffix(".backup")
    if backup.exists():
        backup.unlink()
    shutil.copy2(OUTPUT_PLAN, backup)
    rename_files(rows)
    replace_text_references(rows)
    replace_accessibility_workbook(rows)
    refresh_accessibility_sidecars()
    print("[ok] files renamed and references updated")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
